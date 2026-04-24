"""Excel tool — Create professional .xlsx files via openpyxl.

Produces financial-analyst-quality spreadsheets with formatting, formulas,
conditional formatting, data validation, and smart professional defaults.
"""

import os
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl import load_workbook as _load_workbook

from neo.tools.paths import resolve_path, _validate_write_path

# ---------------------------------------------------------------------------
# Theme definitions
# ---------------------------------------------------------------------------
THEMES: dict[str, dict[str, str]] = {
    "professional": {
        "header_fill": "2B5797",
        "header_font": "FFFFFF",
        "shade_fill": "EDF2F9",
        "accent": "2B5797",
    },
    "minimal": {
        "header_fill": "E0E0E0",
        "header_font": "333333",
        "shade_fill": "F9F9F9",
        "accent": "757575",
    },
    "corporate": {
        "header_fill": "1B1B1B",
        "header_font": "FFFFFF",
        "shade_fill": "F5F5F5",
        "accent": "1B1B1B",
    },
    "colorful": {
        "header_fill": "2E7D32",
        "header_font": "FFFFFF",
        "shade_fill": "E8F5E9",
        "accent": "FF6F00",
    },
}

_DEFAULT_FONT_NAME = "Calibri"
_DEFAULT_FONT_SIZE = 11

# Border styles
_THIN_SIDE = Side(style="thin", color="D0D0D0")
_MEDIUM_SIDE = Side(style="thin", color="999999")

BORDER_PRESETS: dict[str, Border] = {
    "none": Border(),
    "thin": Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE),
    "medium": Border(
        left=_MEDIUM_SIDE,
        right=_MEDIUM_SIDE,
        top=_MEDIUM_SIDE,
        bottom=_MEDIUM_SIDE,
    ),
    "all": Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    ),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _detect_column_type(values: list) -> str:
    """Detect dominant data type in a column from its values."""
    if not values:
        return "text"

    numeric_count = 0
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, (int, float)):
            numeric_count += 1
        elif isinstance(v, str):
            cleaned = v.strip().replace(",", "").replace("$", "").replace("€", "").replace("R$", "")
            try:
                float(cleaned)
                numeric_count += 1
            except (ValueError, TypeError):
                pass

    non_empty = sum(1 for v in values if v is not None and v != "")
    if non_empty == 0:
        return "text"

    return "numeric" if numeric_count / non_empty > 0.5 else "text"


def _get_theme_config(formatting: dict | None) -> dict[str, str]:
    """Resolve theme configuration from formatting dict."""
    if not formatting:
        return THEMES["professional"]

    theme_name = formatting.get("theme", "professional")
    config = dict(THEMES.get(theme_name, THEMES["professional"]))

    # Allow overrides
    if "header_color" in formatting:
        config["header_fill"] = formatting["header_color"].lstrip("#")
    if "header_font_color" in formatting:
        config["header_font"] = formatting["header_font_color"].lstrip("#")

    return config


def _apply_formulas(ws, formulas: list[dict]) -> None:
    """Write formulas to specific cells."""
    for f in formulas:
        cell_ref = f.get("cell", "")
        formula = f.get("formula", "")
        if cell_ref and formula:
            ws[cell_ref] = formula


def _apply_column_formats(
    ws,
    column_formats: dict[str, str],
    headers: list[str],
    data_row_count: int,
) -> None:
    """Apply number formats to columns by header name."""
    header_to_col: dict[str, int] = {}
    for idx, h in enumerate(headers, 1):
        header_to_col[h] = idx

    for header_name, fmt in column_formats.items():
        col_idx = header_to_col.get(header_name)
        if col_idx is None:
            continue

        for row_idx in range(2, data_row_count + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = fmt


def _apply_conditional_formatting(
    ws,
    rules: list[dict],
    headers: list[str],
    data_row_count: int,
) -> None:
    """Apply conditional formatting rules."""
    header_to_col: dict[str, int] = {}
    for idx, h in enumerate(headers, 1):
        header_to_col[h] = idx

    for rule in rules:
        rule_type = rule.get("type", "")
        cell_range = rule.get("range", "")

        # If range uses a header name like "Cost", resolve to column letter range
        if not cell_range and "column" in rule:
            col_idx = header_to_col.get(rule["column"])
            if col_idx:
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{data_row_count + 1}"

        if not cell_range:
            continue

        if rule_type == "color_scale":
            min_color = rule.get("min_color", "F8696B")
            mid_color = rule.get("mid_color", "FFEB84")
            max_color = rule.get("max_color", "63BE7B")
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min",
                    start_color=min_color,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=mid_color,
                    end_type="max",
                    end_color=max_color,
                ),
            )
        elif rule_type == "data_bar":
            color = rule.get("color", "638EC6")
            ws.conditional_formatting.add(
                cell_range,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=color,
                ),
            )
        elif rule_type == "highlight":
            operator = rule.get("operator", "greaterThan")
            value = rule.get("value", 0)
            color = rule.get("color", "FFFF00")
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator=operator,
                    formula=[str(value)],
                    fill=PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid",
                    ),
                ),
            )


def _apply_data_validation(
    ws,
    rules: list[dict],
    headers: list[str],
    data_row_count: int,
) -> None:
    """Apply data validation (dropdown lists, etc.) to columns."""
    header_to_col: dict[str, int] = {}
    for idx, h in enumerate(headers, 1):
        header_to_col[h] = idx

    for rule in rules:
        col_name = rule.get("column", "")
        col_idx = header_to_col.get(col_name)
        if col_idx is None:
            continue

        val_type = rule.get("type", "list")
        col_letter = get_column_letter(col_idx)
        # Apply to data rows (skip header)
        cell_range = f"{col_letter}2:{col_letter}{max(data_row_count + 1, 1000)}"

        if val_type == "list":
            values = rule.get("values", [])
            if values:
                formula = '"' + ",".join(str(v) for v in values) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.error = f"Please select from: {', '.join(str(v) for v in values)}"
                dv.errorTitle = "Invalid Value"
                dv.prompt = f"Select: {', '.join(str(v) for v in values)}"
                dv.promptTitle = col_name
                ws.add_data_validation(dv)
                dv.add(cell_range)


def _auto_fit_columns(ws, headers: list[str], rows: list[list]) -> None:
    """Improved column width auto-fit accounting for data types and formatting."""
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8

        max_len = header_len
        for row in rows:
            if col_idx - 1 < len(row):
                cell_len = len(str(row[col_idx - 1])) if row[col_idx - 1] is not None else 0
                max_len = max(max_len, cell_len)

        # Add padding, enforce min/max
        width = max(max_len + 3, 10)
        width = min(width, 55)

        # Extra space for formatted numbers (currency symbols, separators)
        col_values = [r[col_idx - 1] for r in rows if col_idx - 1 < len(r)]
        if _detect_column_type(col_values) == "numeric":
            width = max(width, 14)

        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------------
def create_workbook(
    title: str,
    sheets: list[dict] | None = None,
    formatting: dict | None = None,
    output_path: str | None = None,
) -> str:
    """Create a professional Excel workbook with formatting, formulas, and validation.

    Args:
        title: Filename (without extension) or full path.
        sheets: List of sheet dicts. Each sheet supports:
            - name (str): Sheet tab name (max 31 chars)
            - headers (list[str]): Column header names
            - rows (list[list]): Data rows (array of arrays)
            - formulas (list[dict]): [{cell: "B10", formula: "=SUM(B2:B9)"}]
            - column_formats (dict): {header_name: excel_format_string}
            - data_validation (list[dict]): [{column, type, values}]
            - conditional_formatting (list[dict]): [{type, range/column, ...}]
            - auto_filter (bool): Enable filter arrows (default True)
        formatting: Document-level style settings:
            - theme (str): "professional", "minimal", "corporate", "colorful"
            - alternate_row_shading (bool): Zebra stripes (default True)
            - borders (str): "none", "thin" (default), "medium", "all"
            - header_color (str): Hex color override for header fill
            - header_font_color (str): Hex color override for header font
            - font (str): Default font name (default "Calibri")
            - font_size (int): Default font size (default 11)
        output_path: Directory where the file should be saved.

    Returns:
        Absolute path to the created .xlsx file.
    """
    wb = Workbook()
    fmt = formatting or {}
    theme = _get_theme_config(fmt)
    font_name = fmt.get("font", _DEFAULT_FONT_NAME)
    font_size = fmt.get("font_size", _DEFAULT_FONT_SIZE)
    use_shading = fmt.get("alternate_row_shading", True)
    border_style = fmt.get("borders", "thin")
    border = BORDER_PRESETS.get(border_style, BORDER_PRESETS["thin"])

    # Build header styles from theme
    header_font = Font(
        name=font_name,
        size=font_size,
        bold=True,
        color=theme["header_font"],
    )
    header_fill = PatternFill(
        start_color=theme["header_fill"],
        end_color=theme["header_fill"],
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    shade_fill = PatternFill(
        start_color=theme["shade_fill"],
        end_color=theme["shade_fill"],
        fill_type="solid",
    )

    if not sheets:
        ws = wb.active
        sheet_name = os.path.basename(title).replace("/", "_").replace("\\", "_")
        ws.title = sheet_name[:31]
    else:
        wb.remove(wb.active)

        for sheet_def in sheets:
            ws = wb.create_sheet(title=sheet_def.get("name", "Sheet")[:31])
            headers = sheet_def.get("headers", [])
            rows = sheet_def.get("rows", [])
            data_row_count = len(rows)

            # --- Write headers with professional styling ---
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                # Freeze header row
                ws.freeze_panes = "A2"

            # --- Write data rows with formatting ---
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(name=font_name, size=font_size)
                    cell.border = border

                    # Auto-detect alignment
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    # Zebra striping
                    if use_shading and row_idx % 2 == 0:
                        cell.fill = shade_fill

            # --- Auto-filter (default on) ---
            enable_filter = sheet_def.get("auto_filter", True)
            if headers and enable_filter:
                last_col = get_column_letter(len(headers))
                last_row = data_row_count + 1
                ws.auto_filter.ref = f"A1:{last_col}{last_row}"

            # --- Column auto-fit ---
            _auto_fit_columns(ws, headers, rows)

            # --- Apply column number formats ---
            col_formats = sheet_def.get("column_formats")
            if col_formats and headers:
                _apply_column_formats(ws, col_formats, headers, data_row_count)

            # --- Apply formulas ---
            formulas = sheet_def.get("formulas")
            if formulas:
                _apply_formulas(ws, formulas)
                # Style formula cells (bold + top border to look like totals)
                for f in formulas:
                    cell_ref = f.get("cell", "")
                    if cell_ref:
                        cell = ws[cell_ref]
                        cell.font = Font(name=font_name, size=font_size, bold=True)
                        cell.border = Border(
                            top=Side(style="double", color=theme["accent"]),
                            bottom=Side(style="thin", color=theme["accent"]),
                            left=border.left,
                            right=border.right,
                        )
                        # Inherit the column's number format if set
                        if col_formats:
                            col_letter = "".join(c for c in cell_ref if c.isalpha())
                            for col_idx_f, h in enumerate(headers, 1):
                                if get_column_letter(col_idx_f) == col_letter and h in col_formats:
                                    cell.number_format = col_formats[h]

            # --- Apply conditional formatting ---
            cond_fmt = sheet_def.get("conditional_formatting")
            if cond_fmt and headers:
                _apply_conditional_formatting(ws, cond_fmt, headers, data_row_count)

            # --- Apply data validation ---
            data_val = sheet_def.get("data_validation")
            if data_val and headers:
                _apply_data_validation(ws, data_val, headers, data_row_count)

            # --- Print setup ---
            ws.print_title_rows = "1:1"  # Repeat header row on every printed page
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # 0 = auto height

    # Save
    file_path = resolve_path(title, ".xlsx", output_dir=output_path)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    return file_path


# ---------------------------------------------------------------------------
# Collaborative editing
# ---------------------------------------------------------------------------
def edit_workbook(
    file_path: str,
    updates: list[dict] | None = None,
    add_rows: list[dict] | None = None,
    delete_rows: list[dict] | None = None,
    add_sheets: list[str] | None = None,
    delete_sheets: list[str] | None = None,
    save_as: str | None = None,
) -> str:
    """Edit an existing Excel workbook — update cells, add/delete rows, manage sheets.

    Args:
        file_path: Path to the existing .xlsx file.
        updates: List of cell updates. Each dict:
            - cell (str): Cell reference like "B5" or "Sheet1!B5"
            - value: New value for the cell
            - format (str, optional): Number format string
            - bold (bool, optional): Make cell bold
        add_rows: List of row additions. Each dict:
            - sheet (str, optional): Sheet name (default: active sheet)
            - values (list): Row values to append
        delete_rows: List of row deletions. Each dict:
            - sheet (str, optional): Sheet name (default: active sheet)
            - row (int): Row number to delete (1-based)
        add_sheets: List of sheet names to create.
        delete_sheets: List of sheet names to delete.
        save_as: Optional alternate save path. If omitted, overwrites original.

    Returns:
        Absolute path to the saved file with a summary of changes.
    """
    from neo.tools.file_reader import _validate_read_path

    real_path = _validate_read_path(file_path)
    wb = _load_workbook(real_path)

    changes: list[str] = []

    # --- Add sheets ---
    if add_sheets:
        for name in add_sheets:
            if name not in wb.sheetnames:
                wb.create_sheet(title=name[:31])
                changes.append(f"Added sheet '{name}'")

    # --- Delete sheets ---
    if delete_sheets:
        for name in delete_sheets:
            if name in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb[name]
                changes.append(f"Deleted sheet '{name}'")

    # --- Update cells ---
    if updates:
        for upd in updates:
            cell_ref = upd.get("cell", "")
            if not cell_ref:
                continue

            # Parse optional sheet prefix: "Sheet1!B5"
            if "!" in cell_ref:
                sheet_name, cell_ref = cell_ref.split("!", 1)
                ws = wb[sheet_name]
            else:
                ws = wb.active

            cell = ws[cell_ref]
            cell.value = upd.get("value")

            if upd.get("format"):
                cell.number_format = upd["format"]
            if upd.get("bold") is not None:
                cell.font = Font(
                    name=cell.font.name or "Calibri",
                    size=cell.font.size or 11,
                    bold=upd["bold"],
                )
            changes.append(f"Updated {cell_ref} = {upd.get('value')}")

    # --- Add rows ---
    if add_rows:
        for entry in add_rows:
            sheet_name = entry.get("sheet")
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            values = entry.get("values", [])
            ws.append(values)
            changes.append(f"Appended row to '{ws.title}'")

    # --- Delete rows (process in reverse to keep indices stable) ---
    if delete_rows:
        # Group by sheet and sort descending
        by_sheet: dict[str, list[int]] = {}
        for entry in delete_rows:
            sheet_name = entry.get("sheet") or wb.active.title
            row_num = entry.get("row", 0)
            if row_num > 0:
                by_sheet.setdefault(sheet_name, []).append(row_num)

        for sheet_name, rows in by_sheet.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_num in sorted(rows, reverse=True):
                    if 1 <= row_num <= ws.max_row:
                        ws.delete_rows(row_num)
                        changes.append(f"Deleted row {row_num} from '{sheet_name}'")

    # --- Save ---
    if save_as:
        out_path = os.path.expanduser(save_as)
        if not out_path.endswith(".xlsx"):
            out_path += ".xlsx"
        _validate_write_path(out_path)
    else:
        out_path = real_path

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

    summary = "; ".join(changes) if changes else "No changes applied"
    return f"Saved {out_path} ({summary})"
